import pdb
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.table import Table, TableStyleInfo
filepath = '..\\Source\\New Source Files\\'

source_filepath = filepath + 'BCData_LGeo_June_24_2024.xlsx'
demand_factor_filepath = filepath + 'Demand_Buffer_Clean_24_06_24.xlsx'
# pdb.set_trace()
output_filepath = os.path.join(os.getcwd(), 'inputs\\BCData_LGeo_June2024_Master_20240708.xlsx')

census_2006_df = pd.read_excel(source_filepath, sheet_name='Census of Population_2006', header=[0, 1, 2])
census_2011_df = pd.read_excel(source_filepath, sheet_name='Census of Population_2011', header=[0, 1, 2])
census_2016_df = pd.read_excel(source_filepath, sheet_name='Census of Population_2016', header=[0, 1, 2])
census_2021_df = pd.read_excel(source_filepath, sheet_name='Census of Population_2021', header=[0, 1, 2])
peh_df = pd.read_excel(source_filepath, sheet_name='PEH_2021_HB')
projections_df = pd.read_excel(source_filepath, sheet_name='Projections_2021-2041')
vacancy_df = pd.read_excel(source_filepath, sheet_name='VacancyRate_2021')
demand_factor_df = pd.read_excel(demand_factor_filepath, sheet_name='DBuff')


def rename_columns(df):
    new_columns = []
    for col in df.columns:
        new_col = tuple('' if 'Unnamed' in str(part) else part for part in col)
        new_columns.append(new_col)
    df.columns = pd.MultiIndex.from_tuples(new_columns)
    return df


def to_multi_level_columns(df):
    multi_level_columns = [('', col, '') for col in df.columns]
    df.columns = pd.MultiIndex.from_tuples(multi_level_columns)
    return df


census_2006_df = rename_columns(census_2006_df)
census_2011_df = rename_columns(census_2011_df)
census_2016_df = rename_columns(census_2016_df)
census_2021_df = rename_columns(census_2021_df)

peh_df = to_multi_level_columns(peh_df)
projections_df = to_multi_level_columns(projections_df)
vacancy_df = to_multi_level_columns(vacancy_df)
demand_factor_df = to_multi_level_columns(demand_factor_df)


merge_col = ('', 'GEOUID', '')
# pdb.set_trace()
merged_df = census_2006_df.merge(census_2011_df, how='left', on=[merge_col], suffixes=('', '_2011')) \
                        .merge(census_2016_df, how='left', on=[merge_col], suffixes=('', '_2016')) \
                        .merge(census_2021_df, how='left', on=[merge_col], suffixes=('', '_2021')) \
                        .merge(demand_factor_df, how='left', right_on=[('', 'GEO_ID', '')], left_on=[merge_col], suffixes=('', '_df'))
# merged_df[('', 'CDID', '')] = pd.to_numeric(merged_df[merge_col].apply(lambda x: str(x)[:4] if len(str(x)) >= 4 else ""), errors='coerce')

if len(merged_df) > len(census_2021_df):
    print("Double check the Geouids, the rows are duplicated...")
elif len(merged_df) < len(census_2021_df):
    print("Double check the join, some rows are dropped...")
else:
    print("The join was successfull...")

# merged_df_peh = merged_df.merge(peh_df[['GeoUID', 'Number of People who Experienced Homelessness, 2021']],
#                                 how='left', left_on=[('', 'CDID', '')], right_on=[('', 'GeoUID', '')])

all_df = merged_df.merge(peh_df, how='left', left_on=[('', 'CD_ID', '')], right_on=[('', 'GeoUID', '')]).merge(
    projections_df, left_on=[merge_col], right_on=[('', 'GeoUID', '')]).merge(vacancy_df, left_on=[merge_col],
                                                                              right_on=[('', 'GeoUID', '')])
all_df[('_x', 'Municipality', '_x')] = all_df[('_x', 'Municipality', '_x')].fillna(all_df[('', 'Geography Name', '')])
all_df[('_x', 'Regional District', '_x')] = all_df[('_x', 'Regional District', '_x')].fillna(all_df[('', 'Geography Name', '')])
all_df[('_x', 'CD_ID', '_x')] = all_df[('_x', 'CD_ID', '_x')].fillna(all_df[('', 'GEOUID', '')].astype(str).str[:4]).astype(int)
# all_df.to_csv(r"L:\Projects\22005 - Housing Needs Assessment\Processed\HART_2024\Archive\test_df.csv")

# pdb.set_trace()

columns_to_remove = [('_x', 'GEOUID', '_x'), ('_x', 'NameTypeUID', '_x'), ('_x', 'Name', '_x'), ('_x', 'Type', '_x'),
                     ('_2021', 'NameTypeUID', '_2021'), ('_2021', 'Name', '_2021'), ('_2021', 'Type', '_2021'),
                     ('_2011', 'NameTypeUID', '_2011'), ('_2011', 'Name', '_2011'), ('_2011', 'Type', '_2011'),
                     ('_2016', 'NameTypeUID', '_2016'), ('_2016', 'Name', '_2016'), ('_2016', 'Type', '_2016'),
                     ('_x', 'GEO_ID', '_x'), ('_x', 'Geography Name', '_x'),
                     ('_x', 'GeoUID', '_x'), ('_x', 'Geography Type', '_x'),
                     ('_y', 'GeoUID', '_y'), ('_y', 'Geography Name', '_y'),
                     ('_y', 'Geography Type', '_y'), ('_y', 'GeoUID', '_y'),
                     ('', 'GeoUID', ''), ('', 'Geography Name', ''), ('', 'Geography Type', '')
                     ]
# pdb.set_trace()

def remove_suffixes(columns):
    new_columns = []
    for col in columns:
        new_col = tuple('' if part in ('_y', '_x') else part for part in col)
        new_columns.append(new_col)
    return pd.MultiIndex.from_tuples(new_columns)


final_df = all_df.loc[:, ~all_df.columns.isin(columns_to_remove)]
final_df.columns = remove_suffixes(final_df.columns)
final_df[('', 'Demand Factor', '')] = final_df[('', 'Demand Factor', '')].fillna(0)

homelessness_column = ('', 'Number of People who Experienced Homelessness, 2021', '')
final_df['Homelessness_Filled'] = final_df[homelessness_column]
# Group by CD_ID and fill missing values
final_df['Homelessness_Filled'] = final_df.groupby([('', 'CD_ID', '')])['Homelessness_Filled'].transform(lambda x: x.fillna(method='ffill').fillna(method='bfill'))
final_df[homelessness_column] = final_df['Homelessness_Filled'].fillna(0)
final_df.drop(columns=['Homelessness_Filled'], inplace=True)

final_df.iloc[:, 326] = final_df.iloc[:, 326].replace('No data', 0).replace('X', 0)

municipalities = ['Nanaimo', 'Powell River', 'Northern Rockies']

# Define the conditions and apply the replacement
final_df[('', 'Municipality', '')] = final_df.apply(
    lambda row: row[('', 'Municipality', '')] + ' ' if row[('', 'Municipality', '')] in municipalities and len(str(row[('', 'GEOUID', '')])) == 4 else row[('', 'Municipality', '')], axis=1
)

# Arranging the dataframe in desired column sequence
final_df = final_df.iloc[:, np.r_[0, 327, 329:331, 1:327, 328, 331:336]]

final_df.to_excel(output_filepath)

wb = load_workbook(output_filepath)
ws = wb.active

ws.cell(row=2, column=1, value="index")
merged_cell_ranges = list(ws.merged_cells.ranges)

for merged_cell_range in merged_cell_ranges:
    ws.unmerge_cells(str(merged_cell_range))
# Adjust the headers to repeat level 0 headers
# Adjust the headers to repeat level 0 headers
new_columns = pd.MultiIndex.from_tuples([('', 'index', '')] + [tuple(col) for col in final_df.columns])

# Add the new column to the DataFrame
final_df = pd.concat([pd.DataFrame({new_columns[0]: range(len(final_df))}), final_df], axis=1)
final_df.columns = new_columns

for col_num, col in enumerate(final_df.columns, 1):
    ws.cell(row=1, column=col_num, value=col[0] if col[0] else "")
    ws.cell(row=2, column=col_num, value=col[1] if col[1] else "")
    ws.cell(row=3, column=col_num, value=col[2] if col[2] else "")

# Save the workbook
wb.save(output_filepath)


id_file = demand_factor_df.drop(columns=[('', 'Demand Factor', '')], axis=1)
id_file.to_excel(os.path.join(os.getcwd(), 'inputs\\Id_file.xlsx'))
